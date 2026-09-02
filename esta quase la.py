import flet as ft
import json
import os
from datetime import datetime
from openpyxl import Workbook

ARQUIVO_ESTOQUE = "meu_estoque.json"
ARQUIVO_VENDAS = "historico_vendas.json"

produtos_padrao = {
    "Coxinha": 1.00,
    "Coca-Cola 1 litro": 4.50,
    "Fanta 1 litro": 4.50,
    "Enroladinho": 1.00,
    "Pastel de queijo": 1.00,
    "Pastel de calabresa": 1.00,
    "Açaí 300 ml": 7.00,
    "Açaí 500 ml": 18.00,
    "Lasanha de frango": 15.00,
    "Escondidinho de carne de sol": 18.00
}

# --- Inicialização Segura do Estoque ---
if os.path.exists(ARQUIVO_ESTOQUE):
    try:
        with open(ARQUIVO_ESTOQUE, "r", encoding="utf-8") as arquivo:
            estoque = json.load(arquivo)
    except (json.JSONDecodeError, KeyError):
        estoque = {}
else:
    estoque = {}

for produto in produtos_padrao:
    if produto not in estoque:
        estoque[produto] = 0

# --- Inicialização Segura do Histórico de Vendas ---
if os.path.exists(ARQUIVO_VENDAS):
    try:
        with open(ARQUIVO_VENDAS, "r", encoding="utf-8") as arquivo:
            historico_vendas = json.load(arquivo)
    except (json.JSONDecodeError, KeyError):
        historico_vendas = []
else:
    historico_vendas = []


def salvar_estoque():
    with open(ARQUIVO_ESTOQUE, "w", encoding="utf-8") as arquivo:
        json.dump(estoque, arquivo, indent=4, ensure_ascii=False)


def salvar_venda(produto, quantidade, total):
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    venda_registro = {
        "data_hora": data_hora,
        "produto": produto,
        "quantidade": quantidade,
        "total": total
    }
    historico_vendas.append(venda_registro)
    with open(ARQUIVO_VENDAS, "w", encoding="utf-8") as arquivo:
        json.dump(historico_vendas, arquivo, indent=4, ensure_ascii=False)


def gerar_excel():
    wb = Workbook()
    
    # Aba 1: Situação do Estoque Atual
    ws_estoque = wb.active
    ws_estoque.title = "Estoque Atual"
    ws_estoque.append(["Produto", "Preço Unitário (R$)", "Quantidade em Estoque"])
    for produto, preco in produtos_padrao.items():
        ws_estoque.append([produto, preco, estoque[produto]])
        
    # Aba 2: Histórico de Vendas Detalhado
    ws_vendas = wb.create_sheet(title="Histórico de Vendas")
    ws_vendas.append(["Data e Hora", "Produto", "Quantidade Vendida", "Total Faturado (R$)"])
    for venda in historico_vendas:
        ws_vendas.append([venda["data_hora"], venda["produto"], venda["quantidade"], venda["total"]])
        
    # Aba 3: Faturamento Total por Data
    ws_datas = wb.create_sheet(title="Faturamento por Data")
    ws_datas.append(["Data", "Faturamento Total do Dia (R$)"])
    
    faturamento_por_dia = {}
    for venda in historico_vendas:
        data_apenas = venda["data_hora"].split(" ")[0]
        faturamento_por_dia[data_apenas] = faturamento_por_dia.get(data_apenas, 0.0) + float(venda["total"])
    
    for data, total_dia in sorted(faturamento_por_dia.items()):
        ws_datas.append([data, total_dia])
        
    wb.save("relatorio_vendas_estoque.xlsx")


def main(page: ft.Page):
    page.title = "Controle de Vendas"
    page.window.width = 390
    page.window.height = 750
    page.scroll = ft.ScrollMode.ADAPTIVE
    page.theme_mode = ft.ThemeMode.LIGHT 

    campo_quantidade = ft.TextField(
        label="Digite a quantidade",
        value="1", 
        keyboard_type=ft.KeyboardType.NUMBER,
        text_align=ft.TextAlign.CENTER,
        width=150
    )

    textos_estoque_tela = {}

    # Função de alerta corrigida e compatível com renderização forçada
    def mostrar_alerta(titulo, mensagem, cor_titulo):
        def fechar_alerta(e):
            alerta.open = False
            page.update()

        alerta = ft.AlertDialog(
            title=ft.Text(titulo, color=cor_titulo, weight=ft.FontWeight.BOLD),
            content=ft.Text(mensagem),
            actions=[ft.TextButton("OK", on_click=fechar_alerta)],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        # Adiciona no overlay da página e força a abertura visual instantânea
        page.overlay.append(alerta)
        alerta.open = True
        page.update()

    def acao_exportar_excel(e):
        try:
            gerar_excel()
            mostrar_alerta("Sucesso ✅", "Planilha 'relatorio_vendas_estoque.xlsx' gerada com sucesso!", ft.Colors.GREEN)
        except Exception as erro:
            mostrar_alerta("Erro ao Exportar ❌", f"Feche o arquivo Excel se ele estiver aberto.\nErro: {erro}", ft.Colors.RED)

    def adicionar_estoque(produto_nome):
        try:
            if not campo_quantidade.value or not campo_quantidade.value.isdigit():
                raise ValueError

            quantidade = int(campo_quantidade.value)
            if quantidade <= 0:
                raise ValueError

            estoque[produto_nome] += quantidade
            salvar_estoque()
            
            textos_estoque_tela[produto_nome].value = f"Estoque: {estoque[produto_nome]}"
            campo_quantidade.value = "1"
            page.update()

            mostrar_alerta("Estoque Atualizado ➕", f"Adicionado: {quantidade} unidade(s)\nProduto: {produto_nome}", ft.Colors.BLUE)

        except ValueError:
            mostrar_alerta("Erro de Quantidade ❌", "Por favor, digite um número inteiro maior que zero.", ft.Colors.RED)

    def vender_produto(produto_nome):
        try:
            if not campo_quantidade.value or not campo_quantidade.value.isdigit():
                raise ValueError

            quantidade = int(campo_quantidade.value)
            if quantidade <= 0:
                raise ValueError

            if quantidade > estoque[produto_nome]:
                mostrar_alerta("Estoque Insuficiente ⚠️", f"Você tentou vender {quantidade}, mas só existem {estoque[produto_nome]} no estoque.", ft.Colors.ORANGE)
                return

            estoque[produto_nome] -= quantidade
            total = quantidade * produtos_padrao[produto_nome] 

            salvar_estoque()
            salvar_venda(produto_nome, quantidade, total)
            
            textos_estoque_tela[produto_nome].value = f"Estoque: {estoque[produto_nome]}"
            campo_quantidade.value = "1"
            page.update()

            mostrar_alerta("Venda Registrada 🎉", f"Produto: {produto_nome}\nQuantidade: {quantidade}x\nTotal: R$ {total:.2f}", ft.Colors.GREEN)

        except ValueError:
            mostrar_alerta("Erro de Quantidade ❌", "Por favor, digite um número inteiro maior que zero.", ft.Colors.RED)

    conteudo_pagina = [
        ft.Row([ft.Text("CONTROLE DE VENDAS", size=24, weight=ft.FontWeight.BOLD)], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row([campo_quantidade], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row([
            ft.ElevatedButton(
                "EXPORTAR PARA EXCEL",
                icon=ft.Icons.GRID_ON,
                color=ft.Colors.WHITE,
                bgcolor=ft.Colors.GREEN_700,
                on_click=acao_exportar_excel,
                width=250
            )
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Divider()
    ]

    for produto, preco in produtos_padrao.items():
        txt_est = ft.Text(f"Estoque: {estoque[produto]}", size=14, weight=ft.FontWeight.W_500)
        textos_estoque_tela[produto] = txt_est

        linha_produto = ft.Container(
            content=ft.Column([
                ft.Text(f"{produto} - R$ {preco:.2f}", size=16, weight=ft.FontWeight.BOLD),
                ft.Row([
                    txt_est,
                    ft.Row([
                        ft.ElevatedButton(
                            "VENDER", 
                            color=ft.Colors.WHITE,
                            bgcolor=ft.Colors.RED_ACCENT,
                            on_click=lambda e, p=produto: vender_produto(p)
                        ),
                        ft.IconButton(
                            icon=ft.Icons.ADD, 
                            icon_color=ft.Colors.BLUE,
                            tooltip="Adicionar Estoque",
                            on_click=lambda e, p=produto: adicionar_estoque(p)
                        )
                    ], spacing=5)
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
            ]),
            padding=10,
            border=ft.Border.all(1, ft.Colors.BLACK12), 
            border_radius=8,
            margin=ft.Margin.only(bottom=10) 
        )
        conteudo_pagina.append(linha_produto)

    page.add(ft.Column(conteudo_pagina))

if __name__ == "__main__":
    ft.app(target=main)
