import flet as ft
import json
import os
from datetime import datetime
from openpyxl import Workbook

ARQUIVO_ESTOQUE = "meu_estoque.json"
ARQUIVO_VENDAS = "historico_vendas.json"

produtos_padrao_iniciais = {
    "Coxinha": 1.00, "Coca-Cola 1 litro": 4.50, "Fanta 1 litro": 4.50,
    "Enroladinho": 1.00, "Pastel de queijo": 1.00, "Pastel de calabresa": 1.00,
    "Açaí 300 ml": 7.00, "Açaí 500 ml": 18.00, "Lasanha de frango": 15.00,
    "Escondidinho de carne de sol": 18.00
}

if os.path.exists(ARQUIVO_ESTOQUE):
    try:
        with open(ARQUIVO_ESTOQUE, "r", encoding="utf-8") as arquivo:
            dados_salvos = json.load(arquivo)
            if dados_salvos and "produtos" in dados_salvos:
                produtos_config = dados_salvos["produtos"]
                estoque = dados_salvos["estoque"]
            else:
                produtos_config = produtos_padrao_iniciais.copy()
                estoque = dados_salvos
    except:
        produtos_config = produtos_padrao_iniciais.copy()
        estoque = {}
else:
    produtos_config = produtos_padrao_iniciais.copy()
    estoque = {}

for prod in produtos_config:
    if prod not in estoque: estoque[prod] = 0

if os.path.exists(ARQUIVO_VENDAS):
    try:
        with open(ARQUIVO_VENDAS, "r", encoding="utf-8") as arquivo:
            historico_vendas = json.load(arquivo)
    except: historico_vendas = []
else:
    historico_vendas = []

def salvar_estoque_e_precos():
    estrutura = {"produtos": produtos_config, "estoque": estoque}
    with open(ARQUIVO_ESTOQUE, "w", encoding="utf-8") as arquivo:
        json.dump(estrutura, arquivo, indent=4, ensure_ascii=False)

def salvar_venda(produto, quantity, total):
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    venda_registro = {"data_hora": data_hora, "produto": produto, "quantidade": quantity, "total": total}
    historico_vendas.append(venda_registro)
    with open(ARQUIVO_VENDAS, "w", encoding="utf-8") as arquivo:
        json.dump(historico_vendas, arquivo, indent=4, ensure_ascii=False)

def gerar_excel():
    wb = Workbook()
    ws_estoque = wb.active
    ws_estoque.title = "Estoque Atual"
    ws_estoque.append(["Produto", "Preço Unitário (R$)", "Quantidade em Estoque"])
    for produto, preco in produtos_config.items():
        ws_estoque.append([produto, preco, estoque[produto]])
        
    ws_vendas = wb.create_sheet(title="Histórico de Vendas")
    ws_vendas.append(["Data e Hora", "Produto", "Quantidade Vendida", "Total Faturado (R$)"])
    for venda in historico_vendas:
        ws_vendas.append([venda["data_hora"], venda["produto"], venda["quantidade"], venda["total"]])
        
    ws_datas = wb.create_sheet(title="Faturamento por Data")
    ws_datas.append(["Data", "Faturamento Total do Dia (R$)"])
    
    faturamento_por_dia = {}
    for venda in historico_vendas:
        partes = venda["data_hora"].split(" ")
        data_apenas = partes if partes else "Sem Data"
        faturamento_por_dia[data_apenas] = faturamento_por_dia.get(data_apenas, 0.0) + float(venda["total"])
    
    for data, total_dia in sorted(faturamento_por_dia.items()):
        ws_datas.append([data, total_dia])
    wb.save("relatorio_vendas_estoque.xlsx")

def main(page: ft.Page):
    page.title = "Controle de Vendas"
    page.window.width = 410
    page.window.height = 780
    page.scroll = ft.ScrollMode.ADAPTIVE
    page.theme_mode = ft.ThemeMode.LIGHT 

    campo_quantidade = ft.TextField(label="Quantidade para Venda/Estoque", value="1", keyboard_type=ft.KeyboardType.NUMBER, text_align=ft.TextAlign.CENTER, width=220)
    
    # CORREÇÃO AQUI: Alterado de size para text_size para rodar sem erros
    campo_novo_nome = ft.TextField(label="Nome do Produto", width=180, text_size=14)
    campo_novo_preco = ft.TextField(label="Preço (Ex: 4.50)", width=120, keyboard_type=ft.KeyboardType.NUMBER, text_size=14)

    textos_estoque_tela = {}
    lista_produtos_container = ft.Column()

    def mostrar_alerta(titulo, mensagem, cor_titulo):
        def fechar_alerta(e):
            alerta.open = False
            page.update()
        alerta = ft.AlertDialog(title=ft.Text(titulo, color=cor_titulo, weight=ft.FontWeight.BOLD), content=ft.Text(mensagem), actions=[ft.TextButton("OK", on_click=fechar_alerta)], actions_alignment=ft.MainAxisAlignment.END)
        page.overlay.append(alerta)
        alerta.open = True
        page.update()

    def acao_exportar_excel(e):
        try:
            gerar_excel()
            mostrar_alerta("Sucesso ✅", "Planilha gerada com sucesso!", ft.Colors.GREEN)
        except Exception as erro:
            mostrar_alerta("Erro ao Exportar ❌", f"Feche o arquivo Excel.\nErro: {erro}", ft.Colors.RED)

    def adicionar_estoque(produto_nome):
        try:
            if not campo_quantidade.value or not campo_quantidade.value.isdigit(): raise ValueError
            quantidade = int(campo_quantidade.value)
            if quantidade <= 0: raise ValueError
            estoque[produto_nome] += quantidade
            salvar_estoque_e_precos()
            textos_estoque_tela[produto_nome].value = f"Estoque: {estoque[produto_nome]}"
            campo_quantidade.value = "1"
            page.update()
            mostrar_alerta("Estoque Atualizado ➕", f"Adicionado: {quantidade} unidade(s)\nProduto: {produto_nome}", ft.Colors.BLUE)
        except ValueError:
            mostrar_alerta("Erro de Quantidade ❌", "Digite um número válido.", ft.Colors.RED)

    def vender_produto(produto_nome):
        try:
            if not campo_quantidade.value or not campo_quantidade.value.isdigit(): raise ValueError
            quantidade = int(campo_quantidade.value)
            if quantidade <= 0: raise ValueError
            if quantidade > estoque[produto_nome]:
                mostrar_alerta("Estoque Insuficiente ⚠️", f"Estoque disponível: {estoque[produto_nome]}", ft.Colors.ORANGE)
                return
            estoque[produto_nome] -= quantidade
            total = quantidade * produtos_config[produto_nome]
            salvar_estoque_e_precos()
            salvar_venda(produto_nome, quantity:=quantidade, total)
            textos_estoque_tela[produto_nome].value = f"Estoque: {estoque[produto_nome]}"
            campo_quantidade.value = "1"
            page.update()
            mostrar_alerta("Venda Registrada 🎉", f"Produto: {produto_nome}\nQuantidade: {quantity}x\nTotal: R$ {total:.2f}", ft.Colors.GREEN)
        except ValueError:
            mostrar_alerta("Erro de Quantidade ❌", "Digite um número válido.", ft.Colors.RED)

    def criar_linha_produto(p_nome, p_preco):
        txt_est = ft.Text(f"Estoque: {estoque[p_nome]}", size=14, weight=ft.FontWeight.W_500)
        textos_estoque_tela[p_nome] = txt_est
        return ft.Container(content=ft.Column([ft.Text(f"{p_nome} - R$ {p_preco:.2f}", size=16, weight=ft.FontWeight.BOLD), ft.Row([txt_est, ft.Row([ft.Button("VENDER", color=ft.Colors.WHITE, bgcolor=ft.Colors.RED_ACCENT, on_click=lambda e, p=p_nome: vender_produto(p)), ft.IconButton(icon=ft.Icons.ADD, icon_color=ft.Colors.BLUE, tooltip="Adicionar Estoque", on_click=lambda e, p=p_nome: adicionar_estoque(p))], spacing=5)], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)]), padding=10, border=ft.Border.all(1, ft.Colors.BLACK12), border_radius=8, margin=ft.Margin.only(bottom=10))

    def cadastrar_novo_produto(e):
        nome = campo_novo_nome.value.strip()
        preco_str = campo_novo_preco.value.strip().replace(",", ".")
        if not nome:
            mostrar_alerta("Erro ❌", "O nome não pode ficar vazio.", ft.Colors.RED)
            return
        if nome in produtos_config:
            mostrar_alerta("Erro ❌", "Produto já cadastrado.", ft.Colors.RED)
            return
        try:
            preco = float(preco_str)
            if preco < 0: raise ValueError
        except ValueError:
            mostrar_alerta("Erro ❌", "Digite um preço válido.", ft.Colors.RED)
            return
        produtos_config[nome] = preco
        estoque[nome] = 0
        salvar_estoque_e_precos()
        lista_produtos_container.controls.append(criar_linha_produto(nome, preco))
        campo_novo_nome.value = ""
        campo_novo_preco.value = ""
        page.update()
        mostrar_alerta("Sucesso 🎉", f"'{nome}' cadastrado com sucesso!", ft.Colors.GREEN)

    for produto, preco in produtos_config.items():
        lista_produtos_container.controls.append(criar_linha_produto(produto, preco))

    page.add(ft.Column([
        ft.Row([ft.Text("CONTROLE DE VENDAS", size=24, weight=ft.FontWeight.BOLD)], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row([campo_quantidade], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row([ft.Button("EXPORTAR PARA EXCEL", icon=ft.Icons.GRID_ON, color=ft.Colors.WHITE, bgcolor=ft.Colors.GREEN_700, on_click=acao_exportar_excel, width=250)], alignment=ft.MainAxisAlignment.CENTER),
        ft.Divider(),
        ft.Container(content=ft.Column([
            ft.Text("Cadastrar Novo Item no Cardápio:", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_GREY_700),
            ft.Row([campo_novo_nome, campo_novo_preco], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            ft.Row([ft.Button("CADASTRAR PRODUTO", icon=ft.Icons.PLAYLIST_ADD, color=ft.Colors.WHITE, bgcolor=ft.Colors.BLUE_700, on_click=cadastrar_novo_produto, width=220)], alignment=ft.MainAxisAlignment.CENTER)
        ]), padding=12, bgcolor=ft.Colors.BLUE_GREY_50, border_radius=8, margin=ft.Margin.only(bottom=15)),
        ft.Text("Lista de Itens:", size=16, weight=ft.FontWeight.BOLD),
        lista_produtos_container
    ]))

if __name__ == "__main__":
    ft.run(main)
